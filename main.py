import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.workbook import Workbook

wb1 = load_workbook("1.xlsx")
wb2 = load_workbook("2.xlsx")
wb3 = load_workbook("3.xlsx")
overall = Workbook()

ws1 = wb1.active
ws2 = wb2.active
ws3 = wb3.active
ws4 = overall.active

for row in ws1.iter_rows(values_only=True):
    a = row
for row in ws2.iter_rows(values_only=True):
    b = row
for row in ws3.iter_rows(values_only=True):
    c = row

res = [a, b, c]
res[0], res[-1] = res[-1], res[0]

for row in res:
    ws4.append(row)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for row in ws4["A1:D3"]:
    for cell in row:
        cell.font = Font(name="Calibri", size=12, bold=True)
        cell.border = thin_border


overall.save('res.xlsx')